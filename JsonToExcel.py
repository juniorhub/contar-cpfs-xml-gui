import json
import os
import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import traceback
from openpyxl.utils import get_column_letter

# Suprimir o aviso de depreciação do Tk
os.environ["TK_SILENCE_DEPRECATION"] = "1"

def flatten_json_to_column(data, parent_key='', sep='_'):
    """
    Flatten nested JSON structure into a list of (key, value) pairs for column-based output.
    """
    items = []
    if isinstance(data, dict):
        for key, value in data.items():
            new_key = f"{parent_key}{sep}{key}" if parent_key else key
            if isinstance(value, (dict, list)):
                items.extend(flatten_json_to_column(value, new_key, sep))
            else:
                items.append((new_key, str(value)))
    elif isinstance(data, list):
        for i, item in enumerate(data):
            new_key = f"{parent_key}_{i}" if parent_key else str(i)
            if isinstance(item, (dict, list)):
                items.extend(flatten_json_to_column(item, new_key, sep))
            else:
                items.append((new_key, str(item)))
    return items

def process_json(json_file, output_dir):
    try:
        if not os.path.exists(json_file):
            return f"Erro: O arquivo '{json_file}' não foi encontrado.", []

        # Lê o arquivo JSON
        with open(json_file, 'r', encoding='utf-8') as f:
            data = json.load(f)

        # Verifica se o JSON é um dicionário
        if not isinstance(data, dict):
            return "Erro: Esperado um objeto JSON no nível superior.", []

        # Inicializa o escritor do Excel
        output_file = os.path.join(output_dir, "resultado_json.xlsx")
        writer = pd.ExcelWriter(output_file, engine='openpyxl')
        records = []

        # 1. Tabela: Dados Emissor (top-level fields, exceto 'mdes_config' subestruturas específicas)
        emissor_data = {k: v for k, v in data.items() if k not in ['mdes_config']}
        if 'mdes_config' in data:
            # Inclui apenas mdes_requestor_config, predigitization_keys, customer_service_keys
            for subkey in ['mdes_requestor_config', 'predigitization_keys', 'customer_service_keys']:
                if subkey in data['mdes_config']:
                    emissor_data[subkey] = data['mdes_config'][subkey]
        if emissor_data:
            pairs = flatten_json_to_column(emissor_data)
            if pairs:
                keys, values = zip(*pairs)
                df_emissor = pd.DataFrame({
                    "Chave": keys,
                    "Valor": values
                })
                df_emissor.to_excel(writer, sheet_name="Dados Emissor", index=False)
                records.append(df_emissor)

        # 2. Tabela: Metodos de Autenticação
        if 'mdes_config' in data and 'activation_methods' in data['mdes_config']:
            activation_methods = data['mdes_config']['activation_methods']
            pairs = flatten_json_to_column(activation_methods)
            if pairs:
                keys, values = zip(*pairs)
                df_methods = pd.DataFrame({
                    "Chave": keys,
                    "Valor": values
                })
                df_methods.to_excel(writer, sheet_name="Metodos de Autenticação", index=False)
                records.append(df_methods)

        # 3. Tabela: Configurações Bins
        if 'mdes_config' in data and 'mdes_product_config' in data['mdes_config']:
            bins_data = []
            for bin_item in data['mdes_config']['mdes_product_config']:
                bin_number = bin_item.get('bin')
                for req_config in bin_item.get('mdes_product_requestor_config', []):
                    requestor_id = req_config.get('requestor_id')
                    for range_config in req_config.get('range_config', []):
                        bins_data.append({
                            'Bin': bin_number,
                            'requestor_id': requestor_id,
                            'range_inicial': range_config.get('range_inicial'),
                            'range_final': range_config.get('range_final'),
                            'product_config_id': range_config.get('product_config_id'),
                            'mobile_app_id': range_config.get('mobile_app_id')
                        })
            if bins_data:
                df_bins = pd.DataFrame(bins_data)
                df_bins.to_excel(writer, sheet_name="Configurações Bins", index=False)
                records.append(df_bins)

        # 4. Tabela: Configurações Arte
        if 'mdes_config' in data and 'mdes_product_art' in data['mdes_config']:
            art_data = []
            for art_item in data['mdes_config']['mdes_product_art']:
                product_id = art_item.get('product_id')
                default_config_id = art_item.get('default_config_id')
                for art_config in art_item.get('art_config', []):
                    art_data.append({
                        'product_id': product_id,
                        'default_config_id': default_config_id,
                        'image_id': art_config.get('image_id'),
                        'product_config_id': art_config.get('product_config_id')
                    })
            if art_data:
                df_art = pd.DataFrame(art_data)
                df_art.to_excel(writer, sheet_name="Configurações Arte", index=False)
                records.append(df_art)

        # Salva o arquivo Excel
        writer.close()

        # Ajusta a largura das colunas para melhor visualização
        from openpyxl import load_workbook
        wb = load_workbook(output_file)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = adjusted_width
        wb.save(output_file)

        # Prepara mensagem de resultado
        mensagem = f"Arquivo JSON processado com sucesso.\n"
        mensagem += f"Total de tabelas geradas: {len(records)}\n"
        mensagem += "Tabelas criadas:\n"
        if 'Dados Emissor' in wb.sheetnames:
            mensagem += "- Dados Emissor: Chaves e valores do emissor\n"
        if 'Metodos de Autenticação' in wb.sheetnames:
            mensagem += "- Metodos de Autenticação: Configurações de autenticação\n"
        if 'Configurações Bins' in wb.sheetnames:
            mensagem += "- Configurações Bins: Lista de bins com requestor_id e configurações de range\n"
        if 'Configurações Arte' in wb.sheetnames:
            mensagem += "- Configurações Arte: Lista de configurações de arte com product_id e image_id\n"
        mensagem += f"Arquivo Excel gerado: {output_file}\n"
        mensagem += "Os dados de Configurações Bins e Configurações Arte foram organizados em formato de tabela."

        return mensagem, records

    except FileNotFoundError:
        return f"Erro: O arquivo '{json_file}' não foi encontrado.", []
    except json.JSONDecodeError:
        return "Erro: Arquivo JSON inválido ou mal formatado.", []
    except Exception as e:
        return f"Erro inesperado: {str(e)}", []

class App:
    def __init__(self, root):
        try:
            self.root = root
            self.root.title("Extrator de JSON para Excel")
            self.root.geometry("600x400")
            
            # Frame principal
            self.frame = ttk.Frame(self.root, padding="10")
            self.frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
            
            # Botão para selecionar arquivo JSON
            self.btn_select = ttk.Button(self.frame, text="Selecionar Arquivo JSON", command=self.select_file)
            self.btn_select.grid(row=0, column=0, columnspan=2, pady=10)
            
            # Label para mostrar o arquivo selecionado
            self.file_label = ttk.Label(self.frame, text="Nenhum arquivo selecionado")
            self.file_label.grid(row=1, column=0, columnspan=2, pady=5)
            
            # Botão para processar o arquivo
            self.btn_process = ttk.Button(self.frame, text="Processar", command=self.process_file, state="disabled")
            self.btn_process.grid(row=2, column=0, columnspan=2, pady=10)
            
            # Área de texto para resultados
            self.result_text = tk.Text(self.frame, height=15, width=60)
            self.result_text.grid(row=3, column=0, columnspan=2, pady=10)
            
            # Variável para armazenar o caminho do arquivo JSON
            self.json_file = ""
        except Exception as e:
            print(f"Erro ao inicializar a interface: {e}")
            traceback.print_exc()
            raise

    def select_file(self):
        try:
            file_path = filedialog.askopenfilename(filetypes=[("Arquivos JSON", "*.json")])
            if file_path:
                self.json_file = file_path
                self.file_label.config(text=f"Arquivo: {os.path.basename(file_path)}")
                self.btn_process.config(state="normal")
        except Exception as e:
            print(f"Erro ao selecionar arquivo: {e}")
            traceback.print_exc()

    def process_file(self):
        try:
            if not self.json_file:
                messagebox.showerror("Erro", "Nenhum arquivo selecionado.")
                return
            
            # Pede ao usuário para selecionar o diretório de saída
            output_dir = filedialog.askdirectory(title="Selecionar Diretório para Salvar o Excel")
            if not output_dir:
                messagebox.showwarning("Aviso", "Nenhum diretório selecionado. O Excel não será gerado.")
                return
            
            # Processa o arquivo
            mensagem, records = process_json(self.json_file, output_dir)
            
            # Atualiza a área de texto com os resultados
            self.result_text.delete(1.0, tk.END)
            self.result_text.insert(tk.END, mensagem)
            
            # Mostra mensagem de sucesso ou erro
            if records:
                messagebox.showinfo("Sucesso", "Processamento concluído. Verifique o arquivo Excel gerado.")
            else:
                messagebox.showerror("Erro", mensagem)
        except Exception as e:
            print(f"Erro ao processar arquivo: {e}")
            traceback.print_exc()

if __name__ == "__main__":
    try:
        root = tk.Tk()
        app = App(root)
        root.mainloop()
    except Exception as e:
        print(f"Erro ao iniciar o aplicativo: {e}")
        traceback.print_exc()
        input("Pressione Enter para sair...")